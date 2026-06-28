# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX = next(ROOT.glob("*.xlsx"))
OUT = ROOT / "backend" / "src" / "main" / "resources" / "data" / "seed-data.json"

CUSTOMER_SHEETS = ["辦公室", "登記", "停業", "登記(發票轉)", "聯絡處"]


def clean(value: Any) -> str | None:
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group()) if match else None


def cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if len(row) > index else None


def is_customer_row(row: tuple[Any, ...]) -> bool:
    company = clean(cell(row, 1))
    if not company or company in {"公司名稱", "全方位商務中心有限公司"}:
        return False
    return any(clean(cell(row, i)) for i in [2, 3, 4, 5])


def registration_type(sheet_name: str) -> str:
    if sheet_name == "辦公室":
        return "實體辦公室"
    if sheet_name == "聯絡處":
        return "聯絡處"
    if sheet_name == "停業":
        return "停業"
    return "登記"


def customer_status(sheet_name: str) -> str:
    return "停業" if sheet_name == "停業" else "租賃中"


def make_office_no(raw: str | None, sheet_name: str, index: int) -> str:
    if raw:
        return raw
    prefix = {
        "登記": "REG",
        "停業": "CLS",
        "登記(發票轉)": "INV",
        "聯絡處": "CNT",
    }.get(sheet_name, "OFF")
    return f"{prefix}-{index:03d}"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    raw_customers: list[dict[str, Any]] = []
    seen_names: set[str] = set()

    for sheet_name in CUSTOMER_SHEETS:
        ws = wb[sheet_name]
        for r in range(5, ws.max_row + 1, 2):
            row1 = tuple(ws.cell(r, c).value for c in range(1, min(ws.max_column, 20) + 1))
            row2 = tuple(ws.cell(r + 1, c).value for c in range(1, min(ws.max_column, 20) + 1))
            if not is_customer_row(row1):
                continue
            company = clean(cell(row1, 1))
            if not company or company in seen_names:
                continue
            seen_names.add(company)

            monthly = []
            for month, col in enumerate(range(7, 19), start=1):
                marker = clean(cell(row1, col - 1))
                receipt = clean(cell(row2, col - 1))
                if marker or receipt:
                    monthly.append({"month": month, "marker": marker, "receiptNo": receipt})

            raw_customers.append(
                {
                    "sourceSheet": sheet_name,
                    "officeNo": clean(cell(row1, 0)),
                    "companyName": company,
                    "ownerName": clean(cell(row1, 2)),
                    "phone": clean(cell(row1, 3)),
                    "leaseStart": clean(cell(row1, 4)),
                    "deposit": number(cell(row1, 5)),
                    "taxId": clean(cell(row2, 2)),
                    "mobile": clean(cell(row2, 3)),
                    "leaseEnd": clean(cell(row2, 4)),
                    "rent": number(cell(row2, 5)),
                    "note": clean(cell(row1, 18)) or clean(cell(row2, 1)),
                    "changeNote": clean(cell(row1, 19)),
                    "registrationType": registration_type(sheet_name),
                    "status": customer_status(sheet_name),
                    "monthly": monthly,
                }
            )

    branch = {"branchId": 1, "branchName": "敦化館"}
    offices: dict[str, dict[str, Any]] = {}
    customers: list[dict[str, Any]] = []
    contracts: list[dict[str, Any]] = []
    rent_payments: list[dict[str, Any]] = []

    for idx, item in enumerate(raw_customers, start=1):
        office_no = make_office_no(item.get("officeNo"), item["sourceSheet"], idx)
        if office_no not in offices:
            offices[office_no] = {
                "officeId": len(offices) + 1,
                "officeNo": office_no,
                "branchId": branch["branchId"],
                "phone": item.get("phone") or item.get("mobile"),
                "notes": item["sourceSheet"],
            }

        customers.append(
            {
                "customerId": idx,
                "companyName": item["companyName"],
                "taxId": item.get("taxId"),
                "status": item.get("status"),
                "ownerName": item.get("ownerName"),
                "contactPerson": item.get("ownerName"),
                "phone": item.get("phone") or item.get("mobile"),
                "forwardingAddress": None,
                "pettyCash": None,
                "notes": item.get("note") or item.get("changeNote"),
                "registrationType": item.get("registrationType"),
            }
        )

        contract_id = idx
        contracts.append(
            {
                "contractId": contract_id,
                "customerId": idx,
                "officeId": offices[office_no]["officeId"],
                "startDateText": item.get("leaseStart"),
                "endDateText": item.get("leaseEnd"),
                "terminationDateText": item.get("leaseEnd") if item.get("status") == "停業" else None,
                "rent": item.get("rent"),
                "deposit": item.get("deposit"),
                "registrationType": item.get("registrationType"),
                "leaseStatus": "已停業" if item.get("status") == "停業" else "租賃中",
            }
        )

        for monthly in item.get("monthly", []):
            rent_payments.append(
                {
                    "rentPaymentId": len(rent_payments) + 1,
                    "customerId": idx,
                    "contractId": contract_id,
                    "paymentMonth": monthly["month"],
                    "paymentDateText": None,
                    "feeStartDateText": item.get("leaseStart"),
                    "feeEndDateText": item.get("leaseEnd"),
                    "amount": item.get("rent"),
                    "receiptNo": monthly.get("receiptNo"),
                    "note": monthly.get("marker"),
                }
            )

    refunds: list[dict[str, Any]] = []
    if "退款" in wb.sheetnames:
        ws = wb["退款"]
        company_to_customer = {c["companyName"]: c["customerId"] for c in customers}
        for r in range(2, ws.max_row + 1):
            company = clean(ws.cell(r, 3).value)
            reason = clean(ws.cell(r, 4).value)
            amount = number(ws.cell(r, 6).value)
            if not company and not reason and amount is None:
                continue
            refunds.append(
                {
                    "refundId": len(refunds) + 1,
                    "customerId": company_to_customer.get(company),
                    "contractId": None,
                    "companyName": company,
                    "reason": reason,
                    "refundAmount": amount,
                    "note": clean(ws.cell(r, 5).value),
                }
            )

    seed = {
        "branches": [branch],
        "offices": list(offices.values()),
        "customers": customers,
        "contracts": contracts,
        "rentPayments": rent_payments,
        "refunds": refunds,
        "roles": [
            {"rolePermissionId": 1, "roleName": "主管", "scope": "全部分館"},
            {"rolePermissionId": 2, "roleName": "管理人員", "scope": "客戶、合約、租金與退款"},
            {"rolePermissionId": 3, "roleName": "一般人員", "scope": "查詢與新增租金紀錄"},
        ],
        "staff": [
            {"staffId": 1, "staffName": "系統主管", "account": "manager", "rolePermissionId": 1},
            {"staffId": 2, "staffName": "管理人員", "account": "supervisor", "rolePermissionId": 2},
            {"staffId": 3, "staffName": "一般人員", "account": "staff", "rolePermissionId": 3},
        ],
        "salesTargets": [
            {"salesTargetId": 1, "branchId": 1, "month": 1, "category": "登記", "targetCount": 3},
            {"salesTargetId": 2, "branchId": 1, "month": 1, "category": "實體辦公室", "targetCount": 1},
        ],
        "performanceBonuses": [
            {"bonusId": 1, "staffId": 1, "period": "115 Q1", "netCount": 0, "bonusAmount": 0}
        ],
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(seed, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUT}")
    print(
        f"customers={len(customers)} offices={len(offices)} "
        f"contracts={len(contracts)} rentPayments={len(rent_payments)} refunds={len(refunds)}"
    )


if __name__ == "__main__":
    main()
